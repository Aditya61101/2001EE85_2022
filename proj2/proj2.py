from datetime import datetime
start_time = datetime.now()
import streamlit as st
import pandas as pd
import openpyxl
from datetime import datetime
from openpyxl.styles import Color, PatternFill, Font, Border, Side
import base64
import io
import tkinter as tk
from tkinter import filedialog
import glob
import os
from zipfile import ZipFile
from io import BytesIO


from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

LabelOrder = [1,-1,2,-2,3,-3,4,-4]

uploadFile = None
MOD = 0

#Help
def proj_octant_gui():
    st.title ("Project2 Solution")

    folderPath = ""

    option = st.selectbox('How many files you want to convert?',('Single file', 'Complete folder'))

    if option == "Single file":
        global uploadedFile
        uploadedFile = st.file_uploader("Upload File", type=['xlsx'],accept_multiple_files=False,key="fileUploader")
        
        if "folderPath" in st.session_state:
            del st.session_state["folderPath"]

    if option == "Complete folder":
        uploadedFile = None

        # Set up tkinter
        root = tk.Tk()
        root.withdraw()

        # Make folder picker dialog appear on top of other windows
        root.wm_attributes('-topmost', 1)

        # Folder picker button
        st.write('Please select a folder:')
        clicked = st.button('Folder Picker')
        if clicked:
            folderPath = filedialog.askdirectory(master=root)
            st.session_state["folderPath"] = folderPath

    if "folderPath" in st.session_state:
        folderPath = st.session_state["folderPath"]
        dirname = st.text_input('Selected folder:', folderPath)


    global MOD
    MOD = st.number_input('Enter MOD value: ', min_value=1,  value=5000, step=1)
    

    convert, download = st.columns(2)

    with convert:
        conv = st.button("Compute")

        if conv:
            if option == "Single file":
                if not uploadedFile:
                    st.warning("Please upload a xlsx file")
                else:
                    with st.spinner("Please wait..."):
                        fileName = uploadedFile.name.split(".xlsx")[0]
                        outputFileName = startConversion(fileName)

                        with download:
                            with open(outputFileName, 'rb') as my_file:
                                st.download_button(label = 'Download File', data = my_file, file_name = outputFileName, mime = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet') 

            elif option == "Complete folder":
                if "folderPath" in st.session_state:
                    folderPath = st.session_state["folderPath"]

                folder = folderPath.split("/")[-1]

                excel_files = glob.glob(os.path.join(folderPath, "*.xlsx"))

                if len(folder) ==0:
                    st.warning("Please select a folder")
                    return

                if len(excel_files)==0:
                    st.warning("No excel sheet found!!")
                    return

                outputFolderName = getOutputFileName(folder) + ".zip"

                zipObj = ZipFile(outputFolderName, "w")

                with st.spinner("Please wait..."):
                    for i,file in enumerate(excel_files):
                        uploadedFile = file
                        fileName = file.split(".xlsx")[0]
                        fileName = fileName.split("\\")[-1]

                        outputFileName = startConversion(fileName)

                        zipObj.write(outputFileName)


                    zipObj.close()


                    with download:
                        with open(outputFolderName, 'rb') as my_file:
                            st.download_button(label="Download result", data=my_file, file_name=outputFolderName)




def startConversion(fileName):

    df = pd.read_excel(uploadedFile)

    # outputFileName = fileName + "_mod_" + str(MOD) + "_"

    # now = datetime.now()
    # dt_string = now.strftime("%Y-%m-%d-%H-%M-%S")

    # outputFileName += dt_string + ".xlsx"

    outputFileName = getOutputFileName(fileName) + ".xlsx"


    
    outputFile = openpyxl.Workbook()
    currentSheet = outputFile.active
    totalCount = 0

    col = 1

    # Variables to store sum variable
    uSum = 0 
    vSum = 0
    wSum = 0

    for key, value in df.items():
        value = value.tolist()
        totalCount = len(value)

        # key -> 2nd row
        currentSheet.cell(row=2, column=col).value = key

        for r, val in enumerate(value):
            if col==2:
                uSum += val
            elif col==3:
                vSum += val
            elif col==4:
                wSum += val

            currentSheet.cell(row=r+3, column=col).value = val        
        
        
        col +=1

    # Calculating average
    try:
        uAvg = round(uSum/totalCount, 3)
        vAvg = round(vSum/totalCount, 3)
        wAvg = round(wSum/totalCount, 3)
    except ZeroDivisionError:
        print("No input data found!!\nDivision by zero occurred!")
        exit()

    # Setting average values
    try:
        currentSheet.cell(row=3, column=5).value = uAvg
        currentSheet.cell(row=3, column=6).value = vAvg
        currentSheet.cell(row=3, column=7).value = wAvg
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()


    # Processing input
    setProcessedDataWithOctant(uAvg, vAvg, wAvg, currentSheet)


    currentSheet.cell(row=1, column=14).value = "Overall Octant Count"
    currentSheet.cell(row=1, column=24).value = "Rank #1 Should be highligted Yellow"
    currentSheet.cell(row=1, column=35).value = "Overall Transition Count"
    currentSheet.cell(row=1, column=45).value = "Longest Subsquence Length"
    currentSheet.cell(row=1, column=49).value = "Longest Subsquence Length with Range"

    currentSheet.cell(row=2, column=36).value = "To"

    firstRow = ["T","U","V","W","U Avg","V Avg","W Avg","U'=U - U avg","V'=V - V avg","W'=W - W avg","Octant"]
    for i, header in enumerate(firstRow):
        currentSheet.cell(row=2, column=i+1).value = header


    FormatOctantOverallRank(currentSheet, MOD, totalCount)

    FormatCountModWise(currentSheet, MOD, totalCount)

    FormatTransitionOverallCount(currentSheet, totalCount)

    # # Function to add mod wise count of transition
    FormatTransitionCountModWise(currentSheet, MOD, totalCount)

    CheckLongestSubsequence(currentSheet, totalCount)

    outputFile.save(outputFileName)


    data = currentSheet.values
    columns = next(data)[0:]


    df = pd.DataFrame(data, columns=columns)

    # df

    # with download:
    #     towrite = io.BytesIO()
    #     df.to_excel(towrite, encoding='utf-8', index=False, header=True)
    #     towrite.seek(0)  # reset pointer
    #     b64 = base64.b64encode(towrite.read()).decode()  # some strings
    #     linko= f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download={outputFileName}>Download file</a>'
    #     st.markdown(linko, unsafe_allow_html=True)

    return outputFileName





def getOutputFileName(inputFileName):
    outputFileName = inputFileName + "_mod_" + str(MOD) + "_"
    now = datetime.now()
    dt_string = now.strftime("%Y-%m-%d-%H-%M-%S")
    outputFileName += dt_string

    return outputFileName




def InitialiseCount(count):
    for item in LabelOrder:
        count[item] = 0

# Method to initialise dictionary with 0 for "LabelOrder" except 'left'
def InitialiseCountExcept(count, left):
    for item in LabelOrder:
        if(item!=left):
            count[item] = 0



def CheckLongestSubsequence(currentSheet, totalCount):
	# Dictinary to store consecutive sequence count
    count = {}

    # Dictionary to store longest count
    longest = {}

    # Initialing dictionary to 0 for all labels
    InitialiseCount(count)
    InitialiseCount(longest)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, totalCount):
        currRow = i+3
        try:
            curr = int(currentSheet.cell(column=11, row=currRow).value)

            # Comparing current and last value
            if(curr==last):
                count[curr]+=1
                longest[curr] = max(longest[curr], count[curr])
                InitialiseCountExcept(count, curr)
            else:
                count[curr]=1
                longest[curr] = max(longest[curr], count[curr])
                InitialiseCountExcept(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()

        # Updating "last" variable
        last = curr

    # Method to Count longest subsequence frequency
    CheckLongestSubsequenceFrequency(longest, currentSheet, totalCount)


def CheckLongestSubsequenceFrequency(longest, currentSheet, totalCount):
    # Dictinary to store consecutive sequence count
    count = {}

    # Dictinary to store frequency count
    frequency = {}

    # Dictionary to store time range
    timeRange = {}

    for label in LabelOrder:
        timeRange[label] = []

    # Initialing dictionary to 0 for all labels
    InitialiseCount(count)
    InitialiseCount(frequency)

    # Variable to check last value
    last = -10

    # Iterating complete excel sheet
    for i in range(0, totalCount):
        currRow = i+3
        try:
            curr = int(currentSheet.cell(column=11, row=currRow).value)
            
            # Comparing current and last value
            if(curr==last):
                count[curr]+=1
            else:
                count[curr]=1        
                InitialiseCountExcept(count, curr)

            # Upading frequency
            if(count[curr]==longest[curr]):
                frequency[curr]+=1

                # Counting starting and ending time of longest subsequence
                end = float(currentSheet.cell(row=currRow, column=1).value)
                start = 100*end - longest[curr]+1
                start/=100

                # Inserting time interval into map
                timeRange[curr].append([start, end])

                # Reseting count dictionary
                InitialiseCount(count)
            else:
                InitialiseCountExcept(count, curr)
        except FileNotFoundError:
            print("File not found!!")
            exit()
        except ValueError:
            print("File content is invalid!!")
            exit()

        # Updating 'last' variable
        last = curr

    # Setting frequency table into sheet
    CheckFrequency(longest, frequency, currentSheet)

    # Setting time range for longest subsequence
    FormatLongestSubsequenceTime(longest, frequency, timeRange, currentSheet)


# Method to set frequency count to sheet
def CheckFrequency(longest, frequency, currentSheet):
    # Iterating "LabelOrder" and updating sheet
    for i in range(9):
        for j in range(3):
            black = "00000000"
            double = Side(border_style="thin", color=black)
            blackBorder = Border(top=double, left=double, right=double, bottom=double)
            currentSheet.cell(row = 3+i, column = 45+j).border = blackBorder

    currentSheet.cell(row=3, column=45).value= "Octant ##"
    currentSheet.cell(row=3, column=46).value= "Longest Subsquence Length"
    currentSheet.cell(row=3, column=47).value= "Count"

    for i, label in enumerate(LabelOrder):
        currRow = i+3
        try:
            currentSheet.cell(row=currRow+1, column=45).value = label	
            currentSheet.cell(column=46, row=currRow+1).value = longest[label]
            currentSheet.cell(column=47, row=currRow+1).value = frequency[label]
        except FileNotFoundError:
            print("File not found!!")
            exit()

# Method to set time range for longest subsequence
def FormatLongestSubsequenceTime(longest, frequency, timeRange, currentSheet):
    # Naming columns number
    lengthCol = 50
    freqCol = 51
    
    # Initial row, just after the header row
    row = 4

    currentSheet.cell(row=3, column = 49).value = "Octant ###"
    currentSheet.cell(row=3, column = 50).value = "Longest Subsquence Length"
    currentSheet.cell(row=3, column = 51).value = "Count"

    # Iterating all octants 
    for octant in LabelOrder:
        try:
            # Setting octant's longest subsequence and frequency data
            currentSheet.cell(column=49, row=row).value = octant
            currentSheet.cell(column=lengthCol, row=row).value = longest[octant]
            currentSheet.cell(column=freqCol, row=row).value = frequency[octant]
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        try:
            # Setting default labels
            currentSheet.cell(column=49, row=row).value = "Time"
            currentSheet.cell(column=lengthCol, row=row).value = "From"
            currentSheet.cell(column=freqCol, row=row).value = "To"
        except FileNotFoundError:
            print("File not found!!")
            exit()

        row+=1

        # Iterating time range values for each octants
        for timeData in timeRange[octant]:
            try:
                # Setting time interval value
                currentSheet.cell(row=row, column=lengthCol).value = timeData[0]
                currentSheet.cell(row=row, column=freqCol).value = timeData[1]
            except FileNotFoundError:
                print("File not found!!")
                exit()
            row += 1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    blackBorder = Border(top=double, left=double, right=double, bottom=double)

    for i in range(3, row):
        for j in range(49, 52):
            currentSheet.cell(row=i, column = j).border = blackBorder




def FormatTransitionCountModWise(currentSheet, mod, totalCount):
    # Counting partitions w.r.t. mod
    try:
        totalPartition = totalCount//mod
    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    # Checking mod value range
    if(mod<0):
        raise Exception("Mod value should be in range of 1-30000")

    if(totalCount%mod!=0):
        totalPartition +=1

    # Initialising row start for data filling
    rowStart = 16

    # Iterating all partitions 
    for i in range (0,totalPartition):
        # Initialising start and end values
        start = i*mod
        end = min((i+1)*mod-1, totalCount-1)

        # Setting start-end values
        try:
            currentSheet.cell(column=35, row=rowStart-1 + 13*i).value = "Mod Transition Count"
            currentSheet.cell(column=35, row=rowStart + 13*i).value = str(start) + "-" + str(end)
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        # Initialising empty dictionary
        transCount = {}
        for a in range (1,5):
            for b in range(1,5):
                transCount[str(a)+str(b)]=0
                transCount[str(a)+str(-b)]=0
                transCount[str(-a)+str(b)]=0
                transCount[str(-a)+str(-b)]=0
                
        # Counting transition for range [start, end)
        for a in range(start, end+1):
            try:
                curr = currentSheet.cell(column=11, row=a+3).value
                next = currentSheet.cell(column=11, row=a+4).value
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

            # Incrementing count for within range value
            if(next!=None):
                transCount[str(curr) + str(next)]+=1

        # Setting transition counts
        setTransitionCount(rowStart + 13*i, transCount, currentSheet)



def FormatTransitionOverallCount(currentSheet, totalCount):
    # Setting value

    # Initialising empty dictionary
    transCount = {}
    for i in range (1,5):
        for j in range(1,5):
            transCount[str(i)+str(j)]=0
            transCount[str(i)+str(-j)]=0
            transCount[str(-i)+str(j)]=0
            transCount[str(-i)+str(-j)]=0
        
    # Iterating octants values to fill dictionary
    start = 0

    # try and except block for string to int conversion
    try:
        last = int(currentSheet["K3"].value)
    except ValueError:
        print("Sheet input can't be converted to int")
        exit()
    except TypeError:
        print("Sheet doesn't contain integer octant")
        exit()


    while(start<totalCount-1):
        # try and except block for string to int conversion
        try:
            curr = int(currentSheet.cell(row= start+4, column=11).value)
            transCount[str(last) + str(curr)]+=1
            last = curr
        except ValueError:
            print("Sheet input can't be converted to int")
            exit()
        except TypeError:
            print("Sheet doesn't contain integer octant")
            exit()

        start += 1
    
    # Setting transitions counted into sheet
    setTransitionCount(2, transCount, currentSheet)

# Function to set Transition count
def setTransitionCount(row, transCount, currentSheet):
    # Setting hard coded inputs
    try:
        currentSheet.cell(row=row, column=36).value = "To"
        currentSheet.cell(row=row+1, column=35).value = "Octant #"
        currentSheet.cell(row=row+2, column=34).value = "From"

        black = "00000000"
        double = Side(border_style="thin", color=black)
        blackBorder = Border(top=double, left=double, right=double, bottom=double)

        for i in range(35, 44):
            for j in range(row+1, row+1+9):
                currentSheet.cell(row=j, column = i).border = blackBorder


    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    # Setting Labels
    for i, label in enumerate(LabelOrder):
        try:
            currentSheet.cell(row=row+1, column=i+36).value=label
            currentSheet.cell(row=row+i+2, column=35).value=label
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    # Setting data
    for i, l1 in enumerate(LabelOrder):
        maxi = -1

        for j, l2 in enumerate(LabelOrder):
            val = transCount[str(l1)+str(l2)]
            maxi = max(maxi, val)

        yellow = "00FFFF00"
        yellowBackground = PatternFill(start_color=yellow, end_color= yellow, fill_type='solid')

        for j, l2 in enumerate(LabelOrder):
            try:
                currentSheet.cell(row=row+i+2, column=36+j).value = transCount[str(l1)+str(l2)]
                if transCount[str(l1)+str(l2)] == maxi:
                    maxi = -1
                    currentSheet.cell(row=row+i+2, column=36+j).fill = yellowBackground
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()



def FormatCountModWise(currentSheet, mod, totalCount):
	# Initialising empty dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # Variable to store last row
    lastRow = -1

    # Iterating loop to set count dictionary
    start = 0
    while(start<totalCount):
        try:
            count[int(currentSheet.cell(row=start+3, column=11).value)] +=1
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start+=1
    
        try:    
            if(start%mod==0):
                # Setting row data
                try:
                    row = 4 + start//mod
                    lastRow = row
                    currentSheet.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(totalCount, start-1))

                    for i, label in enumerate(LabelOrder):
                        currentSheet.cell(row=row, column=15+i).value = count[label]

                    setRankCount(row,count, currentSheet)
                except FileNotFoundError:
                    print("Output file not found!!")
                    exit()
                except ValueError:
                    print("Row or column values must be at least 1 ")
                    exit()

                # Reset count values
                count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
        except ZeroDivisionError:
            print("Mod can't have 0 value")
            exit()

    try:
        if(start%mod!=0):
            # Setting row data
            try:
                row = 5 + start//mod
                lastRow = row
                currentSheet.cell(row=row, column=14).value = str(start-mod) + "-" + str(min(totalCount, start-1))
                for i, label in enumerate(LabelOrder):
                    currentSheet.cell(row=row, column=15+i).value = count[label]
                
                setRankCount(row,count, currentSheet)
            except FileNotFoundError:
                print("Output file not found!!")
                exit()
            except ValueError:
                print("Row or column values must be at least 1 ")
                exit()

    except ZeroDivisionError:
        print("Mod can't have 0 value")
        exit()

    if(lastRow!=-1):
        setOverallOctantRankMap(lastRow, currentSheet)

def setOverallOctantRankMap(lastRow, currentSheet):
    count = {-1:0, 1:0,  -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # for i in range(5, lastRow+1):
    #     rank1Oct = int(currentSheet.cell(row=i, column=30).value)
    #     count[rank1Oct]+=1
    row =4
    while currentSheet.cell(row=row, column=29).value is not None:
        oct = int(currentSheet.cell(row=row, column=31).value)
        count[oct]+=1
        row+=1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    blackBorder = Border(top=double, left=double, right=double, bottom=double)

    for i in range(9):
        for j in range(3):
            row = lastRow+2+i
            col = 29+j
            currentSheet.cell(row=row, column = col).border = blackBorder


    currentSheet.cell(column=29, row=lastRow+2).value = "Octant ID"
    currentSheet.cell(column=30, row=lastRow+2).value = "Octant Name "
    currentSheet.cell(column=31, row=lastRow+2).value = "Count of Rank 1 Mod Values"

    octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
    for j, oct in enumerate(LabelOrder):
        currentSheet.cell(column=29, row=lastRow+3+j).value = oct
        currentSheet.cell(column=30, row=lastRow+3+j).value = octant_name_id_mapping[oct]
        currentSheet.cell(column=31, row=lastRow+3+j).value = count[oct]




def FormatOctantOverallRank(currentSheet, mod, totalCount):
    firstRow = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    totalRows = totalCount//mod+1+1 # header + overall
    if totalCount%mod!=0:
        totalRows+=1

    black = "00000000"
    double = Side(border_style="thin", color=black)
    blackBorder = Border(top=double, left=double, right=double, bottom=double)

    for i, header in enumerate(firstRow):
        for j in range(totalRows):
            currentSheet.cell(row=3+j, column = 14+i).border = blackBorder


    for i, header in enumerate(firstRow):
        currentSheet.cell(row=3, column = i+14).value = header

    currentSheet.cell(row=4, column = 13).value = "Mod " + str(mod)

    setOverallCount(totalCount, currentSheet)

def setOverallCount(totalCount, currentSheet):	
	# Initialising count dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}

    # Incrementing count dictionary data
    try:
        for i in range (3,totalCount+3):
            count[int(currentSheet.cell(column=11, row=i).value)] = count[int(currentSheet.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()


    # Setting data into sheet
    for i, label in enumerate(LabelOrder):
        try:
            currentSheet.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    setRankCount(4, count, currentSheet)

def setRankCount(row,countMap, currentSheet):
    # Copying the count list to sort
    sortedCount = []
    count = []
    for label in LabelOrder:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
        # for j in range(0, 8):
        #     ell = sortedCount[j]
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break

    rank1Oct = -10

    yellow = "00FFFF00"
    yellowBackground = PatternFill(start_color=yellow, end_color= yellow, fill_type='solid')

    for j in range(0,8):
        currentSheet.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = LabelOrder[j]
            currentSheet.cell(row = row, column=23+j).fill = yellowBackground    

    octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
    currentSheet.cell(row=row , column=31).value = rank1Oct
    currentSheet.cell(row=row , column=32).value = octant_name_id_mapping[rank1Oct]





def setProcessedDataWithOctant(uAvg, vAvg, wAvg, currentSheet):
    start = 3
    time = currentSheet.cell(start, 1).value

    # Iterating throught sheet
    while(time!=None):
        # Calculating processed data
        try:
            u1 = currentSheet.cell(start, 2).value - uAvg
            v1 = currentSheet.cell(start, 3).value - vAvg
            w1 = currentSheet.cell(start, 4).value - wAvg
            
            u1 = round(u1,3)
            v1 = round(v1,3)
            w1 = round(w1,3)

            oct = get_octant(u1, v1, w1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        # Setting processed data
        try:
            currentSheet.cell(row=start, column=8).value = u1
            currentSheet.cell(row=start, column=9).value = v1
            currentSheet.cell(row=start, column=10).value = w1
            currentSheet.cell(row=start, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()


        start = start+1
        try:
            time = currentSheet.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

# Method based on if-else to return octant type
def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2

    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3

    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4





proj_octant_gui()

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))