import openpyxl
wb = openpyxl.load_workbook(r'input_octant_transition_identify.xlsx')
sheet = wb.active
row_count=sheet.max_row
total_count=row_count-1

# List to store octant signs
Octant_Sign_List = [1, -1, 2, -2, 3, -3, 4, -4]

def count_in_range(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    start=2
    fill=4
    while start<total_count:
        # loop to initialize the cell value as 0
        try:
            for j in range(14,22):
                sheet.cell(row=fill,column=j).value=0

            # loop to fill the count of octant in appropriate cell
            for i in range(start,min(row_count+1,mod+start)):
                region = sheet.cell(row=i,column=11).value
                if region==1:
                    sheet.cell(row=fill,column=14).value+=1
                elif region==-1:
                    sheet.cell(row=fill,column=15).value+=1
                elif region==2:
                    sheet.cell(row=fill,column=16).value+=1
                elif region==-2:
                    sheet.cell(row=fill,column=17).value+=1
                elif region==3:
                    sheet.cell(row=fill,column=18).value+=1
                elif region==-3:
                    sheet.cell(row=fill,column=19).value+=1
                elif region==4:
                    sheet.cell(row=fill,column=20).value+=1
                else:
                    sheet.cell(row=fill,column=21).value+=1
            # filling range value in appropriate cell
            x1=str(start-2)
            y1=str(min(total_count-1,mod+start-3))
            sheet.cell(row=fill,column=13).value=x1+'-'+y1

            #increasing initial value for 'for' loop by mod
            start+=mod
            #increasing filling row by 1
            fill+=1
        except FileNotFoundError:
            print('File not found!')
            exit()

def check_octant_sign(u, v, w):
    if u > 0:
        if v > 0:
            if w > 0:
                # this means (u,v) is +ve hence 1st quad and w is +ve so +1
                return 1
            else:
                # this means (u,v) is +ve hence 1st quad but w is -ve so -1
                return -1
        else:
            if w > 0:
                # this means u>0 and v<0 hence 4th quad and w is +ve so +4
                return 4
            else:
                # this means u>0 and v<0 hence 4th quad but w is -ve so -4
                return -4
    else:
        if v > 0:
            if w > 0:
                # this means u<0 and v>0 hence 4th quad and w is +ve so +2
                return 2
            else:
                # this means u<0 and v>0 hence 4th quad but w is -ve so -2
                return -2
        else:
            if w > 0:
                # this means u<0 and v<0 hence 4th quad and w is +ve so +3
                return 3
            else:
                # this means u<0 and v<0 hence 4th quad and w is -ve so -3
                return -3

def avg_calc():
    data_U=0
    data_V=0
    data_W=0
    for i in range(2, row_count + 1):
        try:
            data_U += sheet.cell(row=i,column=2).value
            data_V += sheet.cell(row=i,column=3).value
            data_W += sheet.cell(row=i,column=4).value
        except FileNotFoundError:
            print('File not found!')
    sheet['E1']='u_avg'
    sheet['F1']='v_avg'
    sheet['G1']='w_avg'
    # average of u, v, w
    try:
        u_avg=data_U/total_count
        v_avg=data_V/total_count
        w_avg=data_W/total_count
    except(ZeroDivisionError):
        print("No input data found!!\nDivision by zero is not allowed!")
        exit()

    try:
        #saving average of U in the sheet
        sheet['E2']=u_avg
        #saving average of V in the sheet
        sheet['F2']=v_avg
        #saving average of W in the sheet
        sheet['G2']=w_avg
    except FileNotFoundError:
        print("File not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    sheet['H1']='U-u_avg'
    sheet['I1']='V-v_avg'
    sheet['J1']='W-w_avg'
    for i in range(2, row_count + 1):
        #calculating and saving U-u_avg in the sheet
        sub_u_avg = sheet.cell(row=i,column=2).value-u_avg
        sheet.cell(row=i,column=8).value=sub_u_avg

        #calculating and saving V-v_avg in the sheet
        sub_v_avg = sheet.cell(row=i,column=3).value-v_avg
        sheet.cell(row=i,column=9).value=sub_v_avg

        #calculating and saving W-w_avg in the sheet
        sub_w_avg = sheet.cell(row=i,column=4).value-w_avg
        sheet.cell(row=i,column=10).value=sub_w_avg

def octant_identification(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    #function to calculate and save average value of U, V, W
    avg_calc()
    sheet['K1']='Octant'
    #initializing count values of each octant sign as 0
    for j in range(14,22):
        sheet.cell(row=2,column=j).value=0

    #saving the sign of the octant
    for i in range(2,row_count+1):
        try:
            sub_u_avg=sheet.cell(row=i,column=8).value
            sub_v_avg=sheet.cell(row=i,column=9).value
            sub_w_avg=sheet.cell(row=i,column=10).value
            try:
                octant_sign=check_octant_sign(sub_u_avg,sub_v_avg,sub_w_avg)
            except NameError:
                print('Either the function is not defined or is not named correctly')
                exit()
            if octant_sign==1:
                sheet.cell(row=2,column=14).value+=1
            elif octant_sign==-1:
                sheet.cell(row=2,column=15).value+=1
            elif octant_sign==2:
                sheet.cell(row=2,column=16).value+=1
            elif octant_sign==-2:
                sheet.cell(row=2,column=17).value+=1
            elif octant_sign==3:
                sheet.cell(row=2,column=18).value+=1
            elif octant_sign==-3:
                sheet.cell(row=2,column=19).value+=1
            elif octant_sign==4:
                sheet.cell(row=2,column=20).value+=1
            elif octant_sign==-4:
                sheet.cell(row=2,column=21).value+=1
            sheet.cell(row=i,column=11).value=octant_sign
        except FileNotFoundError:
            print('File not found!')
            exit()

    sheet['L3']='user input'
    sheet['M1']='Octant ID'
    sheet['M2']='Overall Count'
    sheet['M3']= 'Mod '+ str(mod)
    # filling 1, -1, 2, -2, 3, -3, 4, -4 value as a heading of columns
    for j, label in enumerate(Octant_Sign_List):
        sheet.cell(row=1, column=j+14).value=label

    # calling function to count octant sign in the given range
    count_in_range(mod)
    
# Function to set Transition count
def setTransitionCount(row, transition_count):
    try:
        sheet.cell(row=row, column=14).value = "To"
        sheet.cell(row=row+1, column=13).value = "Count"
        sheet.cell(row=row+2, column=12).value = "From"
    except FileNotFoundError:
        print("File not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    # filling Labels
    for i, label in enumerate(Octant_Sign_List):
        try:
            sheet.cell(row=row+1, column=i+14).value=label
            sheet.cell(row=row+i+2, column=13).value=label
        except FileNotFoundError:
            print('File not found!')
            exit()
    # filling data
    for i, a in enumerate(Octant_Sign_List):
        for j, b in enumerate(Octant_Sign_List):
            sheet.cell(row=row+i+2, column=14+j).value = transition_count[str(a)+str(b)]
    
def setOverAllTransitionCount():
    sheet["M13"].value = "Overall Transition Count"
    # Initializing empty dictionary
    transition_count = {}
    for i in range (1,5):
        for j in range(1,5):
            transition_count[str(i)+str(j)]=0
            transition_count[str(i)+str(-j)]=0
            transition_count[str(-i)+str(j)]=0
            transition_count[str(-i)+str(-j)]=0

    # Iterating octants values to fill dictionary
    start = 0
    last_value = sheet.cell(row=2,column=11).value
    while(start<total_count-1):
        curr_value = sheet.cell(row=start+3, column=11).value
        transition_count[str(last_value) + str(curr_value)]+=1
        last_value = curr_value
        start += 1
    # function to set transitions counted into sheet
    try:
        setTransitionCount(14, transition_count)
    except NameError: 
        print('Either the function is not defined or is not named correctly')
        exit()

def setTransitionCountOverMods(mod=5000):
    if mod>30000:
        raise Exception('mod value must be less than or equal to 30000')
    # Counting partitions w.r.t. mod
    total_partitions = total_count//mod
    if(total_count%mod!=0):
        total_partitions +=1

    # Initializing row start for data filling
    row_start = 27

    # Iterating all partitions
    for i in range (0,total_partitions):
        # Initializing start and end values
        start = i*mod
        end = min((i+1)*mod-1, total_count-1)

        # Setting start-end values
        sheet.cell(row=row_start-1 + 13*i, column=13).value = "Mod Transition Count"
        sheet.cell(row=row_start + 13*i , column=13).value = str(start) + "-" + str(end)

        # Initializing empty dictionary
        transition_count = {}
        for x in range (1,5):
            for y in range(1,5):
                transition_count[str(x)+str(y)]=0
                transition_count[str(x)+str(-y)]=0
                transition_count[str(-x)+str(y)]=0
                transition_count[str(-x)+str(-y)]=0
                
        # Counting transition for range [start, end]
        for x in range(start, end+1):
            try:
                curr_value = sheet.cell(column=11, row=x+2).value
                next_value = sheet.cell(column=11, row=x+3).value
                # Incrementing count for within range value
                if(next_value!=None):
                    transition_count[str(curr_value) + str(next_value)]+=1
            except FileNotFoundError:
                print('File not found!')
                exit()
        try:
            # Setting transition count
            setTransitionCount(row_start + 13*i, transition_count)
        except NameError:
            print('Either the function is not defined or the name is not correct')
            exit()

def octant_transition_count(mod):
    # Function to add overall transition count
    try:
        setOverAllTransitionCount()
    except NameError: 
        print('Either the function is not defined or is not named correctly')
        exit()
    # Function to add mod wise count of transition
    try:
        setTransitionCountOverMods(mod)
    except NameError:
        print('Either the function is not defined or is not named correctly')
        exit()

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

mod=5000
try:
    #function to identify octant sign for particular u,v,w values
    octant_identification(mod)
except NameError:
    print('Either the function is not defined or is not named correctly')
    exit()
#function to count transition count of octant sign
try:
    octant_transition_count(mod)
except NameError:
    print('Either the function is not defined or is not named correctly')
    exit()
#saving the file in the given xlsx file
wb.save('output_octant_transition_identify.xlsx')