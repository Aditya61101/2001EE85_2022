import openpyxl
wb = openpyxl.load_workbook(r'input_octant_longest_subsequence.xlsx')
sheet = wb.active
row_count=sheet.max_row
total_count=row_count-1

# List to store octant signs
Octant_Sign_List = [1, -1, 2, -2, 3, -3, 4, -4]

#Help https://youtu.be/H37f_x4wAC0
from datetime import datetime
start_time = datetime.now()

def octant_longest_subsequence_count():
    sheet['M1']='Count'
    sheet['N1']='Longest Subsequence Length'
    sheet['O1']='Count'

    #creating the table for storing longest subsequence length and count
    for i, label in enumerate(Octant_Sign_List):
        sheet.cell(row=i+2,column=13).value=label
    try:
        # loop to calculate max_length and frequency for each octant
        for j, label in enumerate(Octant_Sign_List):
            count_length, max_length, count= 0, -1, 0
            for i in range(2, row_count+1):

                #if the octant sign matches the label value then increase the count_length
                if label == sheet.cell(row=i,column=11).value:
                    count_length+=1

                # if it doesn't then calculate the max length and make the count_length again to 0
                else:
                    if count_length>max_length:
                        count=0
                    # calculating maximum length
                    max_length=max(max_length,count_length)

                    # if the count length again becomes equal to the max_length then this means that there is another occurrence of same length subsequence hence increase the count frequency by 1
                    if max_length==count_length:
                        count+=1

                    # making count_length = 0 again
                    count_length=0
            sheet.cell(row=j+2,column=14).value=max_length
            sheet.cell(row=j+2,column=15).value=count
    except FileNotFoundError:
        print('File not found')
        exit()
def check_octant_sign(u, v, w):
    if u > 0:
        if v > 0:
            if w > 0:
                # this means (u,v) is +ve hence 1st quad and w is +ve so +1
                return 1
            else:
                # this means (u,v) is +ve hence 1st quad but w is -ve so -1
                return -1
        else:
            if w > 0:
                # this means u>0 and v<0 hence 4th quad and w is +ve so +4
                return 4
            else:
                # this means u>0 and v<0 hence 4th quad but w is -ve so -4
                return -4
    else:
        if v > 0:
            if w > 0:
                # this means u<0 and v>0 hence 4th quad and w is +ve so +2
                return 2
            else:
                # this means u<0 and v>0 hence 4th quad but w is -ve so -2
                return -2
        else:
            if w > 0:
                # this means u<0 and v<0 hence 4th quad and w is +ve so +3
                return 3
            else:
                # this means u<0 and v<0 hence 4th quad and w is -ve so -3
                return -3

def avg_calc():
    data_U=0
    data_V=0
    data_W=0
    for i in range(2, row_count + 1):
        try:
            data_U += sheet.cell(row=i,column=2).value
            data_V += sheet.cell(row=i,column=3).value
            data_W += sheet.cell(row=i,column=4).value
        except FileNotFoundError:
            print('File not found!')
    sheet['E1']='u_avg'
    sheet['F1']='v_avg'
    sheet['G1']='w_avg'
    # average of u, v, w
    try:
        u_avg=data_U/total_count
        v_avg=data_V/total_count
        w_avg=data_W/total_count
    except(ZeroDivisionError):
        print("No input data found!!\nDivision by zero is not allowed!")
        exit()

    try:
        #saving average of U in the sheet
        sheet['E2']=u_avg
        #saving average of V in the sheet
        sheet['F2']=v_avg
        #saving average of W in the sheet
        sheet['G2']=w_avg
    except FileNotFoundError:
        print("File not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    sheet['H1']='U-u_avg'
    sheet['I1']='V-v_avg'
    sheet['J1']='W-w_avg'
    for i in range(2, row_count + 1):
        #calculating and saving U-u_avg in the sheet
        sub_u_avg = sheet.cell(row=i,column=2).value-u_avg
        sheet.cell(row=i,column=8).value=sub_u_avg

        #calculating and saving V-v_avg in the sheet
        sub_v_avg = sheet.cell(row=i,column=3).value-v_avg
        sheet.cell(row=i,column=9).value=sub_v_avg

        #calculating and saving W-w_avg in the sheet
        sub_w_avg = sheet.cell(row=i,column=4).value-w_avg
        sheet.cell(row=i,column=10).value=sub_w_avg

def octant_identification():
    #function to calculate and save average value of U, V, W
    avg_calc()
    sheet['K1']='Octant'
    #saving the sign of the octant
    for i in range(2,row_count+1):
        try:
            sub_u_avg=sheet.cell(row=i,column=8).value
            sub_v_avg=sheet.cell(row=i,column=9).value
            sub_w_avg=sheet.cell(row=i,column=10).value
            try:
                octant_sign=check_octant_sign(sub_u_avg,sub_v_avg,sub_w_avg)
            except NameError:
                print('Either the function is not defined or is not named correctly')
                exit()
            sheet.cell(row=i,column=11).value=octant_sign
        except FileNotFoundError:
            print('File not found!')
            exit()

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

try:
    octant_identification()
except NameError:
    print('Either no such function exists or the name is not created')
    exit()
try:
    octant_longest_subsequence_count()
except NameError:
    print('Either no such function exists or the name is not created')
    exit()
#saving the file in the given xlsx file
wb.save('output_octant_longest_subsequence.xlsx')

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))