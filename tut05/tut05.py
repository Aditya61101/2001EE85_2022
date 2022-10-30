import openpyxl
wb = openpyxl.load_workbook(r'octant_input.xlsx')
sheet = wb.active
row_count=sheet.max_row
total_count=row_count-1
from datetime import datetime
start_time = datetime.now()

# List to store octant signs
Octant_Sign_List = [1, -1, 2, -2, 3, -3, 4, -4]

#Help https://youtu.be/N6PBd4XdnEw

###Code
def octant_ranking(mod):
    octant_name_id_mapping = {"1":"Internal outward interaction", "-1":"External outward interaction", "2":"External Ejection", "-2":"Internal Ejection", "3":"External inward interaction", "-3":"Internal inward interaction", "4":"Internal sweep", "-4":"External sweep"}
    dynamic_row = (30000//mod)+8
    for i, label in enumerate(Octant_Sign_List):
        sheet.cell(row=1, column=i+22).value = label
        sheet.cell(row=2, column=i+22).value = 'Rank of ' + str(label)
        sheet.cell(row=dynamic_row+i+1,column=14).value=label
        sheet.cell(row=dynamic_row+i+1,column=15).value=octant_name_id_mapping[str(label)]
    
    sheet.cell(row=2, column=30).value = 'Rank1 Octant ID'
    sheet.cell(row=2, column=31).value = 'Rank1 Octant Name'
    sheet.cell(row=dynamic_row, column=14).value = 'Octant ID'
    sheet.cell(row=dynamic_row, column=15).value = 'Octant Name'
    sheet.cell(row=dynamic_row, column=16).value = 'Count of Rank1 Mod values'

    # octant ranking Code
    for i in range(3,5+(30000//mod)):
        if i==4:
            continue
        list_tuple_rank=[]
        for j in range(14,22):
            count = int(sheet.cell(row=i,column=j).value)
            column_no = j
            list_tuple_rank.append((column_no,count))

        # sorting in descending order on the basis of count values
        list_tuple_rank.sort(key=lambda x:x[1], reverse=True)
        # storing the rank in the sheet
        rank=1
        for pair in list_tuple_rank:
            sheet.cell(row = i,column = pair[0]+8).value=rank
            rank+=1

def count_in_range(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    start=2
    fill=5
    while start<total_count:
        # loop to initialize the cell value as 0
        try:
            for j in range(14,22):
                sheet.cell(row=fill,column=j).value=0

            # loop to fill the count of octant in appropriate cell
            for i in range(start,min(row_count+1,mod+start)):
                region = sheet.cell(row=i,column=11).value
                if region==1:
                    sheet.cell(row=fill,column=14).value+=1
                elif region==-1:
                    sheet.cell(row=fill,column=15).value+=1
                elif region==2:
                    sheet.cell(row=fill,column=16).value+=1
                elif region==-2:
                    sheet.cell(row=fill,column=17).value+=1
                elif region==3:
                    sheet.cell(row=fill,column=18).value+=1
                elif region==-3:
                    sheet.cell(row=fill,column=19).value+=1
                elif region==4:
                    sheet.cell(row=fill,column=20).value+=1
                else:
                    sheet.cell(row=fill,column=21).value+=1
            # filling range value in appropriate cell
            x1=str(start-2)
            y1=str(min(total_count-1,mod+start-3))
            sheet.cell(row=fill,column=13).value=x1+'-'+y1

            #increasing initial value for 'for' loop by mod
            start+=mod
            #increasing filling row by 1
            fill+=1
        except FileNotFoundError:
            print('File not found!')
            exit()

def check_octant_sign(u, v, w):
    if u > 0:
        if v > 0:
            if w > 0:
                # this means (u,v) is +ve hence 1st quad and w is +ve so +1
                return 1
            else:
                # this means (u,v) is +ve hence 1st quad but w is -ve so -1
                return -1
        else:
            if w > 0:
                # this means u>0 and v<0 hence 4th quad and w is +ve so +4
                return 4
            else:
                # this means u>0 and v<0 hence 4th quad but w is -ve so -4
                return -4
    else:
        if v > 0:
            if w > 0:
                # this means u<0 and v>0 hence 4th quad and w is +ve so +2
                return 2
            else:
                # this means u<0 and v>0 hence 4th quad but w is -ve so -2
                return -2
        else:
            if w > 0:
                # this means u<0 and v<0 hence 4th quad and w is +ve so +3
                return 3
            else:
                # this means u<0 and v<0 hence 4th quad and w is -ve so -3
                return -3

def avg_calc():
    data_U=0
    data_V=0
    data_W=0
    for i in range(2, row_count + 1):
        try:
            data_U += sheet.cell(row=i,column=2).value
            data_V += sheet.cell(row=i,column=3).value
            data_W += sheet.cell(row=i,column=4).value
        except FileNotFoundError:
            print('File not found!')
    sheet['E1']='u_avg'
    sheet['F1']='v_avg'
    sheet['G1']='w_avg'
    # average of u, v, w
    try:
        u_avg=data_U/total_count
        v_avg=data_V/total_count
        w_avg=data_W/total_count
    except(ZeroDivisionError):
        print("No input data found!!\nDivision by zero is not allowed!")
        exit()
    try:
        #saving average of U in the sheet
        sheet['E2']=u_avg
        #saving average of V in the sheet
        sheet['F2']=v_avg
        #saving average of W in the sheet
        sheet['G2']=w_avg
    except FileNotFoundError:
        print("File not found!!")
        exit()
    except ValueError:
        print("Row or column values must be at least 1 ")
        exit()

    sheet['H1']='U-u_avg'
    sheet['I1']='V-v_avg'
    sheet['J1']='W-w_avg'
    for i in range(2, row_count + 1):
        #calculating and saving U-u_avg in the sheet
        sub_u_avg = sheet.cell(row=i,column=2).value-u_avg
        sheet.cell(row=i,column=8).value=sub_u_avg

        #calculating and saving V-v_avg in the sheet
        sub_v_avg = sheet.cell(row=i,column=3).value-v_avg
        sheet.cell(row=i,column=9).value=sub_v_avg

        #calculating and saving W-w_avg in the sheet
        sub_w_avg = sheet.cell(row=i,column=4).value-w_avg
        sheet.cell(row=i,column=10).value=sub_w_avg

def octant_identification(mod):
    if mod>30000:
        raise Exception('mod value should be less than or equal to 30000')
    #function to calculate and save average value of U, V, W
    avg_calc()

    sheet['K1']='Octant'
    #initializing count values of each octant sign as 0
    for j in range(14,22):
        sheet.cell(row=3,column=j).value=0
    #saving the sign of the octant
    for i in range(2,row_count+1):
        try:
            sub_u_avg=sheet.cell(row=i,column=8).value
            sub_v_avg=sheet.cell(row=i,column=9).value
            sub_w_avg=sheet.cell(row=i,column=10).value
            try:
                octant_sign=check_octant_sign(sub_u_avg,sub_v_avg,sub_w_avg)
            except NameError:
                print('Either the function is not defined or is not named correctly')
                exit()
            if octant_sign==1:
                sheet.cell(row=3,column=14).value+=1
            elif octant_sign==-1:
                sheet.cell(row=3,column=15).value+=1
            elif octant_sign==2:
                sheet.cell(row=3,column=16).value+=1
            elif octant_sign==-2:
                sheet.cell(row=3,column=17).value+=1
            elif octant_sign==3:
                sheet.cell(row=3,column=18).value+=1
            elif octant_sign==-3:
                sheet.cell(row=3,column=19).value+=1
            elif octant_sign==4:
                sheet.cell(row=3,column=20).value+=1
            elif octant_sign==-4:
                sheet.cell(row=3,column=21).value+=1
            sheet.cell(row=i,column=11).value=octant_sign
        except FileNotFoundError:
            print('File not found!')
            exit()
    sheet['L4']='user input'
    sheet['M2']='Octant ID'
    sheet['M3']='Overall Count'
    sheet['M4']= 'Mod '+ str(mod)
    # filling 1, -1, 2, -2, 3, -3, 4, -4 value as a heading of columns
    for j, label in enumerate(Octant_Sign_List):
        sheet.cell(row=2, column=j+14).value=label

    # calling function to count octant sign in the given range
    count_in_range(mod)

from platform import python_version
ver = python_version()

if ver == "3.8.10":
    print("Correct Version Installed")
else:
    print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")


mod=5000
try:
    octant_identification(mod)
except NameError:
    print('Either the function name is wrong or the function does not exists')
try:
    octant_ranking(mod)
except NameError:
    print('Either the function name is wrong or the function does not exists')

wb.save('octant_output_ranking_excel.xlsx')

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
