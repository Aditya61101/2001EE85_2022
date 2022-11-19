from datetime import datetime
start_time = datetime.now()
import os
os.system("cls")

import openpyxl
import glob

from openpyxl.styles import Color, PatternFill, Font, Border, Side

from platform import python_version
ver = python_version()

if ver == "3.8.10":
	print("Correct Version Installed")
else:
	print("Please install 3.8.10. Instruction are present in the GitHub Repo/Webmail. Url: https://pastebin.com/nvibxmjw")

##Read all the excel files in a batch format from the input/ folder. Only xlsx to be allowed
##Save all the excel files in a the output/ folder. Only xlsx to be allowed
## output filename = input_filename[_octant_analysis_mod_5000].xlsx , ie, append _octant_analysis_mod_5000 to the original filename.

octant_sign = [1,-1,2,-2,3,-3,4,-4]
octant_name_id_mapping = {1:"Internal outward interaction", -1:"External outward interaction", 2:"External Ejection", -2:"Internal Ejection", 3:"External inward interaction", -3:"Internal inward interaction", 4:"Internal sweep", -4:"External sweep"}
yellow = "00FFFF00"
yellow_bg = PatternFill(start_color=yellow, end_color= yellow, fill_type='solid')
black = "00000000"
double = Side(border_style="thin", color=black)
black_border = Border(top=double, left=double, right=double, bottom=double)

###Code
def set_rank_count(row,countMap, outputSheet):
    # Copying the count list to sort
    sortedCount = []
    count = []
    for label in octant_sign:
        count.append(countMap[label])

    for ct in count:
        sortedCount.append(ct)

    sortedCount.sort(reverse=True)

    rank = []

    for i, el in enumerate(count):
        for j, ell in enumerate(sortedCount):
            if(ell==el):
                rank.append(j+1)
                sortedCount[j] = -1
                break
    rank1Oct = -10

    for j in range(0,8):
        outputSheet.cell(row = row, column=23+j).value = rank[j]
        if(rank[j]==1):
            rank1Oct = octant_sign[j]
            outputSheet.cell(row = row, column=23+j).fill = yellow_bg    

    outputSheet.cell(row=row , column=31).value = rank1Oct
    outputSheet.cell(row=row , column=32).value = octant_name_id_mapping[rank1Oct]

def setOverallCount(total_count, outputSheet):	
	# Initializing count dictionary
    count = {-1:0, 1:0, -2:0, 2:0, -3:0, 3:0, -4:0, 4:0}
    # Incrementing count dictionary data
    try:
        for i in range (3,total_count+3):
            count[int(outputSheet.cell(column=11, row=i).value)] = count[int(outputSheet.cell(column=11, row=i).value)] +1
    except FileNotFoundError:
        print("Output file not found!!")
        exit()
    except ValueError:
        print("Sheet input can't be converted to int or row/colum should be atleast 1")
        exit()
    except TypeError:
        print("Sheet doesn't contact valid octant value!!")
        exit()

    # Setting data into sheet
    for i, label in enumerate(octant_sign):
        try:
            outputSheet.cell(row=4, column=i+15).value = count[label]
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

    set_rank_count(4, count, outputSheet)

def set_overall_octant_rank_count(outputSheet, mod, total_count):
    headers = ["Octant ID",1,-1,2,-2,3,-3,+4,-4,"Rank Octant 1", "Rank Octant -1","Rank Octant 2","Rank Octant -2","Rank Octant 3","Rank Octant -3","Rank Octant 4","Rank Octant -4","Rank1 Octant ID","Rank1 Octant Name"]

    totalRows = total_count//mod+1+1 # header + overall
    if total_count%mod!=0:
        totalRows+=1

    for i, header in enumerate(headers):
        for j in range(totalRows):
            outputSheet.cell(row=3+j, column = 14+i).border = black_border

    for i, header in enumerate(headers):
        outputSheet.cell(row=3, column = i+14).value = header

    outputSheet.cell(row=4, column = 13).value = "Mod " + str(mod)

    setOverallCount(total_count, outputSheet)

# Method based on if-else to return octant type
def get_octant(x,y,z):
    if(x>=0 and y>=0):
        if(z>=0):
            return 1
        else:
            return -1
    
    if(x<0 and y>=0):
        if(z>=0):
            return 2
        else:
            return -2

    if(x<0 and y<0):
        if(z>=0):
            return 3
        else:
            return -3

    if(x>=0 and y<0):
        if(z>=0):
            return 4
        else:
            return -4

def setProcessedDataWithOctant(u_avg, v_avg, w_avg, total_count, inputSheet, outputSheet):
    start = 2
    time = inputSheet.cell(start, 1).value

    # Iterating through out the sheet
    while(time!=None):
        # Calculating processed data
        try:
            u1 = inputSheet.cell(start, 2).value - u_avg
            v1 = inputSheet.cell(start, 3).value - v_avg
            w1 = inputSheet.cell(start, 4).value - w_avg
            
            u1 = round(u1,3)
            v1 = round(v1,3)
            w1 = round(w1,3)

            oct = get_octant(u1, v1, w1)
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        # Setting processed data
        try:
            outputSheet.cell(row=start+1, column=8).value = u1
            outputSheet.cell(row=start+1, column=9).value = v1
            outputSheet.cell(row=start+1, column=10).value = w1
            outputSheet.cell(row=start+1, column=11).value = oct
        except FileNotFoundError:
            print("Output file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

        start = start+1
        try:
            time = inputSheet.cell(start, 1).value
        except FileNotFoundError:
            print("Input file not found!!")
            exit()
        except ValueError:
            print("Row or column values must be at least 1 ")
            exit()

def set_input_data(input_file_name, outputSheet):
	input_file = openpyxl.load_workbook(input_file_name)
	inputSheet = input_file.active

	start = 2
	time = inputSheet.cell(start, 1).value

    # Variables to store sum variable
	u_sum = 0 
	v_sum = 0
	w_sum = 0

	# Iterating complete file till time value is not None
	while(time!=None):
		try:
			u_sum += float(inputSheet.cell(start, 2).value)
			v_sum += float(inputSheet.cell(start, 3).value)
			w_sum += float(inputSheet.cell(start, 4).value)
		except ValueError:
			print("Sheet input can't be converted to float!!")
			exit()
		except TypeError:
			print("Sheet doesn't contain valid float input!!")
			exit()

		try:
			# Setting input time,u,v,w values
			outputSheet.cell(row=start+1, column=1).value = inputSheet.cell(start, 1).value 
			outputSheet.cell(row=start+1, column=2).value = inputSheet.cell(start, 2).value 
			outputSheet.cell(row=start+1, column=3).value = inputSheet.cell(start, 3).value 
			outputSheet.cell(row=start+1, column=4).value = inputSheet.cell(start, 4).value 
		except FileNotFoundError:
			print("File not found!!")
			exit()
		except ValueError:
			print("Row or column values must be at least 1 ")
			exit()

		start = start+1
		time = inputSheet.cell(start, 1).value

	# Setting total count
	total_count = start-2 # -1 for header and -1 for last None
	# Calculating average
	try:
		u_avg = round(u_sum/total_count, 3)
		v_avg = round(v_sum/total_count, 3)
		w_avg = round(w_sum/total_count, 3)
	except ZeroDivisionError:
		print("No input data found!!\nDivision by zero occurred!")
		exit()

	# Setting average values
	try:
		outputSheet.cell(row=3, column=5).value = u_avg
		outputSheet.cell(row=3, column=6).value = v_avg
		outputSheet.cell(row=3, column=7).value = w_avg
	except FileNotFoundError:
		print("Output file not found!!")
		exit()
	except ValueError:
		print("Row or column values must be at least 1 ")
		exit()

	# Processing input
	setProcessedDataWithOctant(u_avg, v_avg, w_avg, total_count, inputSheet, outputSheet)

	return total_count

def entry_point(input_file, mod):
	fileName = input_file.split("\\")[-1]
	fileName = fileName.split(".xlsx")[0]
	outputFileName = "output/" + fileName + "_octant_analysis_mod_" + str(mod) + ".xlsx"

	outputFile = openpyxl.Workbook()
	outputSheet = outputFile.active

	outputSheet.cell(row=1, column=14).value = "Overall Octant Count"
	outputSheet.cell(row=1, column=24).value = "Rank #1 Should be highligted Yellow"
	outputSheet.cell(row=1, column=35).value = "Overall Transition Count"
	outputSheet.cell(row=1, column=45).value = "Longest Subsquence Length"
	outputSheet.cell(row=1, column=49).value = "Longest Subsquence Length with Range"
	outputSheet.cell(row=2, column=36).value = "To"

	headers = ["T", "U", "V", "W", "U Avg", "V Avg", "W Avg", "U'=U - U avg", "V'=V - V avg","W'=W - W avg", "Octant"]
	for i, header in enumerate(headers):
		outputSheet.cell(row=2, column=i+1).value = header

	total_count = set_input_data(input_file, outputSheet)
	set_overall_octant_rank_count(outputSheet, mod, total_count)
	outputFile.save(outputFileName)

def octant_analysis(mod=5000):
	path = os.getcwd()
	csv_files = glob.glob(os.path.join(path + "\input", "*.xlsx"))
	
	for file in csv_files:
		entry_point(file, mod)
mod=5000
octant_analysis(mod)

#This shall be the last lines of the code.
end_time = datetime.now()
print('Duration of Program Execution: {}'.format(end_time - start_time))
